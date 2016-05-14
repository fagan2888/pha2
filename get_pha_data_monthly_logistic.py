from __future__ import division
import os, sys
import numpy as np
import scipy.stats
import scipy.cluster.vq
from k_means import KMeans

np.set_printoptions(suppress=True)

try:
    import openpyxl
except ImportError:
    print "This script requires the openpyxl module to access the data in Microsoft Excel files."
    print "Please execute 'sudo pip install openpyxl' or 'pip install openpyxl' (Windows)."
    raise

# get_scenario_data.py should already have been run, creating a folder with standard inputs
inputs_dir = "inputs"
pha_dir = os.path.join(inputs_dir, "pha_125_logistic")
pha_mean_dir = pha_dir + '_mean'
n_scenarios = 1000000

n_digits = 4 # len(str(n_scenarios-1))  # how many leading zeros to use for scenario names
base_year = 2015

# set target rates of price growth (net of inflation)
# numbers below are average real growth rate shown in EIA AEO 2015 forecast for 2016-2040, 
# which is the basis for the HECO 2016-04-01 PSIP forecast.
# See Switch-Hawaii/data/EIA-based fuel cost forecasts/HECO fuel cost forecasts PSIP 2016-04.xlsx
oil_monthly_avg_increase = 0.002388632   # 2.9%/year
gas_monthly_avg_increase = 0.002617286   # 3.2%/year

# date when random forecast begins
fuel_base_date = 2016    # 1/1/16

# fossil fuel base prices, 2016$/mmbtu, from EIA AEO 2015 forecast, also used in HECO 2016-04 PSIP
# (note: these are higher than actual fuel prices in 2016)
# see Switch-Hawaii/data/EIA-based fuel cost forecasts/HECO fuel cost forecasts PSIP 2016-04.xlsx
oil_base_price = 12.264923    # Brent Crude $/bbl / 6.115	
gas_base_price = 3.902859139   # Henry Hub gas

# factors to calculate various fuel prices from base prices; 
# each is a tuple of (oil multiplier, gas multiplier, constant adder).
price_factors = {}

# oil multipliers found by fitting HECO forecast from 2016 PSIP to 
# EIA AEO 2015 forecast of Brent Crude prices, in nominal terms. 
# See "HECO fuel cost forecasts PSIP 2016-04.xlsx". 
# Note: we do the fitting in nominal prices, and then use those
# multipliers and adders when calculating real prices, because HECO
# doesn't seem to have accounted for inflation in their adders (especially
# important for LNG, where the adder is a large part of the price, and HECO
# ends up with LNG rising more slowly than Henry Hub in absolute real dollars.)
# As a result, we get somewhat higher fossil fuel prices than HECO in real terms.
price_factors["LSFO", "base"] = (1.1166, 0.0, -0.0466)
price_factors["Diesel", "base"] = (1.1842, 0.0, 1.8633)
price_factors["ULSD", "base"] = (1.2254, 0.0, 2.486)

# # oil multipliers found by fitting Hawaii-wide LSFO and ULSD prices
# # for 2000-2015 to 3-month lag of Brent Crude, as shown in
# # Ulupono's 'Fuel Costs Monte Carlo-3.xlsx!DBEDT-HEI-BBG Fuel Jan'00-pres'
# price_factors["LSFO", "base"] = (1.135544872, 0.0, -2.582515866/6.115)
# price_factors["Diesel", "base"] = (1.129139616, 0.0, 15.13061141/6.115)

# price for 40% LSFO/60% ULSD shown in 2016-04-01 PSIP p. J-9
# note: this is a more diesel-intense blend than shown in prior PSIP
# (in 2014 PSIP, Diesel/LSFO blend was used to meet MATS emission requirements;
# 2016 PSIP says that's not necessary after all, but 60/40 ULSD/LSFO could be 
# needed to meet NAAQS in the future)
price_factors["LSFO-Diesel-Blend", "base"] = tuple(
    0.4*lsfo + 0.6*ulsd 
        for lsfo, ulsd in zip(price_factors["LSFO", "base"], price_factors["ULSD", "base"])
)
# diesel price + delta based on 2016 biodiesel vs. diesel in 2016-04-01 PSIP p. J-9
price_factors["Biodiesel", "base"] = \
    price_factors["Diesel", "base"][:2] + (price_factors["Diesel", "base"][2] + (32.81-16.29),)

# bulk LNG cost from Ulupono spreadsheet (including liquifaction and delivery but not FRSU)
# note: this seems to be broadly consistent with the containerized price below
price_factors["LNG", "bulk"] = (0.0, 1.2, 6.0)
# containerized LNG cost found by fitting all-in cost of containerized LNG in
# 2016-04-01 PSIP p. J-16 to EIA AEO 2015 Reference nominal price for Henry Hub
# (see "HECO fuel cost forecasts.xlsx")
price_factors["LNG", "container"] = (0.0, 1.1578, 8.0776)



def rand_uniform(size, R):
    # Generates n vector-valued variates with uniform marginals and correlation matrix R.
    # Returns an n x N matrix, where N is the order of R.
    # from http://www.mathworks.com/matlabcentral/newsreader/view_thread/8900
    # which cites "Paul L. Fackler, Modeling Interdependence, AJAE 1991, pp 1091-1097." 
    # and "S.T. Li and J.L. Hammond Generation of Psuedo-random Numbers etc,
    # IEEE Transactions on Systems Man and Cybernetics, Sept. 1975, pp. 557-61."
    # also see http://comisef.wikidot.com/tutorial:correlateduniformvariates 
    if hasattr(R, "shape"):
        if not R.shape:     # shape is an empty tuple, i.e., single numpy number
            R = np.array([[1.0, R], [R, 1.0]])
    else:
        if hasattr(R, "len"):   # standard python list or tuple; pretty unlikely
            R = np.array(R)
        else:   # single number
            R = np.array([[1.0, R], [R, 1.0]])
    #print "R={}".format(R)
    N = R.shape[0]
    # find the covariance matrix for a standard normal distribution that will have
    # rank correlation R
    r = 2.0 * np.sin((np.pi / 6.0) * R)
    return scipy.stats.norm.cdf(np.random.multivariate_normal(np.repeat(0.0, N), r, size))

def trunc_logistic(quantiles, mean, std):
    # return values from a logistic distribution with a particular mean and standard deviation, 
    # selected by quantiles x. Values will be truncated at -0.99 and +0.99
    # See https://en.wikipedia.org/wiki/Logistic_distribution for std <-> loc conversion
    return scipy.stats.logistic.ppf(
        quantiles, loc=mean, scale=std*np.sqrt(3.0)/np.pi
    ).clip(-0.99, 2*mean+0.99)


if not os.path.exists(pha_dir):
    os.makedirs(pha_dir)
if not os.path.exists(pha_mean_dir):
    os.makedirs(pha_mean_dir)

# read standard data
with open(os.path.join(inputs_dir, "fuel_supply_curves.tab")) as f:
    standard_fuel_costs = [r.split("\t") for r in f.read().splitlines()]

headers = standard_fuel_costs.pop(0)
period_ind = headers.index("period")
periods = sorted(set(float(r[period_ind]) for r in standard_fuel_costs))


# build an empirical version of the joint distribution of price changes in 
# oil (lsfo/diesel blend) and LNG, based on behavior in 2000-2015.
# also factor out inflation of about 0.19% per month during this period
# opening the workbook is the slowest part of the whole script
print "loading 'Fuel Costs Monte Carlo-3.xlsx'"
wb = openpyxl.load_workbook("Fuel Costs Monte Carlo-3.xlsx", data_only=True)
ws = wb["DBEDT-HEI-BBG Fuel Jan'00-pres"]
print "reading data from 'Fuel Costs Monte Carlo-3.xlsx'..."
oil_historical_prices = np.array([r[0].value for r in ws["K3:K194"]])
gas_historical_prices = np.array([r[0].value for r in ws["N3:N194"]])
oil_historical_multipliers = (1 - 0.0019) * oil_historical_prices[1:]/oil_historical_prices[:-1]
gas_historical_multipliers = (1 - 0.0019) * gas_historical_prices[1:]/gas_historical_prices[:-1]
print "finished reading from 'Fuel Costs Monte Carlo-3.xlsx'"

# calculate statistics for oil and gas historical changes
oil_historical_changes = oil_historical_multipliers - 1.0
gas_historical_changes = gas_historical_multipliers - 1.0
oil_historical_std = oil_historical_changes.std()
gas_historical_std = gas_historical_changes.std()
oil_gas_rank_corr = scipy.stats.spearmanr(oil_historical_changes, gas_historical_changes).correlation
print "targets: oil_monthly_avg_increase={}, gas_monthly_avg_increase={}".format(
    oil_monthly_avg_increase, gas_monthly_avg_increase)
print "oil average historical change={}, gas average historical change={}".format(
    oil_historical_changes.mean(), gas_historical_changes.mean())
print "oil_historical_std={}, gas_historical_std={}".format(oil_historical_std, gas_historical_std)
print "oil_gas_rank_corr = {}".format(oil_gas_rank_corr)
print "oil - gas historical correlation = {}".format(
    scipy.stats.pearsonr(oil_historical_changes, gas_historical_changes)[0]
)
# NOTE: we have a weird property that the arithmetic mean of changes in gas prices is 
# higher than oil, but the cumulative product of 1+monthly_change leads to an increase
# in oil prices and a decrease in gas prices by the end of the historical record.
# If we ever try to use historical means for future work, maybe they should be geometric
# means of the monthly multipliers? (then subtract 1.0)

# list of all months from the forecast start date till the start of the last period
months = np.arange(fuel_base_date, periods[-1]+1.0/12, 1.0/12)
# indices of the months that are closest to the start of each period.
period_starts = np.array([np.argmin(np.abs(months - p)) for p in periods])

# define a random walk through the future months, drawing from historical variations
# sample_draws = np.random.randint(0, len(oil_historical_multipliers), size=(len(months), n_scenarios))
# oil_multipliers = np.cumprod(oil_historical_multipliers[sample_draws], axis=0)[period_starts, :]
# gas_multipliers = np.cumprod(gas_historical_multipliers[sample_draws], axis=0)[period_starts, :]

n_trajectories = 125
save_vars = ['oil_prices', 'gas_prices', 'mu', 'cluster_id']

if os.path.exists(os.path.join(pha_dir, save_vars[0] + '.npy')):
    # retrieve saved scenario data
    for var in save_vars:
        f = os.path.join(pha_dir, var + '.npy')
        locals()[var] = np.load(f)
else:
    # construct the requested number of scenarios in batches of no more than 100000,
    # to keep memory requirements reasonable
    oil_multipliers = np.zeros((len(period_starts), n_scenarios))
    gas_multipliers = np.zeros((len(period_starts), n_scenarios))
    oil_sim_sum = 0.0
    gas_sim_sum = 0.0
    oil_sim_ssd = 0.0
    gas_sim_ssd = 0.0
    sim_size = 0
    for start in range(0,n_scenarios,100000):
        batch_size = min(100000, n_scenarios-start)
        sys.stdout.write("starting group {}-{}\n".format(start, start+batch_size-1))
        sys.stdout.flush()
        # oil_gas_sim_ranks size = (n_months, batch_size, 2)
        oil_gas_sim_ranks = rand_uniform(size=(len(months), batch_size), R=oil_gas_rank_corr)
        oil_sim_changes = trunc_logistic(
            oil_gas_sim_ranks[:,:,0], mean=oil_monthly_avg_increase, std=oil_historical_std
        )
        gas_sim_changes = trunc_logistic(
            oil_gas_sim_ranks[:,:,1], mean=gas_monthly_avg_increase, std=gas_historical_std
        )
        oil_multipliers[:,start:(start+batch_size)] \
            = np.cumprod(1.0 + oil_sim_changes, axis=0)[period_starts, :]
        gas_multipliers[:,start:(start+batch_size)] \
            = np.cumprod(1.0 + gas_sim_changes, axis=0)[period_starts, :]
        # save data for mean and std dev calculations
        sim_size += oil_sim_changes.size
        oil_sim_sum += oil_sim_changes.sum()
        gas_sim_sum += gas_sim_changes.sum()
        oil_sim_ssd += oil_sim_changes.var() * oil_sim_changes.size
        gas_sim_ssd += gas_sim_changes.var() * gas_sim_changes.size

    # np.set_printoptions(suppress=True, linewidth=150)
    # print "oil_multipliers"
    # print oil_multipliers
    # print "gas_multipliers"
    # print gas_multipliers

    oil_prices = oil_base_price * oil_multipliers
    gas_prices = gas_base_price * gas_multipliers

    print "generated {} price scenarios".format(n_scenarios)
    print "mean values: oil_sim_changes={}, gas_sim_changes={}".format(
        oil_sim_sum/sim_size, gas_sim_sum/sim_size)
    print "standard deviations: oil_sim_changes={}, gas_sim_changes={}".format(
        np.sqrt(oil_sim_ssd/sim_size), np.sqrt(gas_sim_ssd/sim_size))
    # print "rank correlation of changes={}".format(
    #     scipy.stats.spearmanr(oil_sim_changes.ravel(), gas_sim_changes.ravel()).correlation)
    # print "correlation of changes = {}".format(
    #     scipy.stats.pearsonr(oil_sim_changes.ravel(), gas_sim_changes.ravel())[0])
    print "mean prices (oil, gas):"
    print oil_prices.mean(axis=1)
    print gas_prices.mean(axis=1)
    print "annual growth rate of average prices: oil={}, gas={}".format(
        (oil_multipliers[-1,:].mean()/oil_multipliers[0,:].mean()) ** (1/(periods[-1]-periods[0])),
        (gas_multipliers[-1,:].mean()/gas_multipliers[0,:].mean()) ** (1/(periods[-1]-periods[0]))
    )

    # cluster the oil and gas trajectories into 125 representative scenarios
    # (note: our dimensions are oil_2021, oil_2029, oil_2037, oil_2045, 
    # gas_2021, gas_2029, gas_2037 and gas_2045)

    # km = KMeans(125, np.hstack([oil_prices.T, gas_prices.T]))
    # km.init_centers()
    # km.find_centers()
    # # km.plot() # can't plot because it's 8-dimensional
    # oil_price_traj = km.mu[:,:len(periods)]
    # gas_price_traj = km.mu[:,-len(periods):]
    # traj_weight = np.bincount(km.cluster_id)
    # pperiods = np.array(periods)[np.newaxis,:]

    # weight = np.sqrt(traj_weight/traj_weight.max())
    # for i, traj in enumerate(oil_price_traj):
    #     # plot each row as a separate series, with appropriate width and alpha
    # #     plt.semilogy(periods, traj, 'k-', linewidth=5*traj_weight[i]/traj_weight.mean(), alpha=.1)
    #     plt.semilogy(periods, traj, 'k-', linewidth=10*weight[i], alpha=weight[i])
    # plt.show()

    # mu, cluster_id = scipy.cluster.vq.kmeans2(
    #     data=np.hstack([oil_prices.T, gas_prices.T]),
    #     k=125,
    #     minit='points'
    # )

    # get a better starting point than scipy kmeans usually provides
    km = KMeans(125, np.hstack([oil_prices.T, gas_prices.T]))
    km.init_centers()   # takes about 60 s for 100,000; roughly linear in #
    mu, cluster_id = scipy.cluster.vq.kmeans2(
        data=np.hstack([oil_prices.T, gas_prices.T]), 
        k=km.mu
    )

    for var in save_vars:
        f = os.path.join(pha_dir, var + '.npy')
        np.save(f, locals()[var])

# process the scenario data

oil_price_traj = mu[:,:len(periods)]    # first half of mu
gas_price_traj = mu[:,-len(periods):]   # second half of mu
traj_weight = np.bincount(cluster_id)/cluster_id.shape[0]
# print traj_weight

# # calculate quantiles for analysis
# mu_sort_idx = np.argsort(mu, axis=0)
# mu_col_idx = np.arange(mu.shape[1])
# mu_sort = mu[mu_sort_idx, mu_col_idx]
# mu_rank = np.cumsum(traj_weight[mu_sort_idx], axis=0) - 0.5 * traj_weight[mu_sort_idx]
# for rank in [0.1, 0.25, 0.5, 0.75, 0.9]:
#     # plot quantiles or save in a file...
#     for c in mu.shape[1]:
#         r_price = np.interp(x=rank, xp=mu_rank[:,c], fp=mu_sort[:,c])

def plot():
    import matplotlib.pyplot as plt
    from mpl_toolkits.mplot3d import Axes3D     # magic to activate 3d axes

    # show all the clusters
    fig = plt.figure()
    ax = fig.gca()
    for i, traj in enumerate(oil_price_traj):
        # plot each row as a separate series, with appropriate width and alpha
        ax.plot(periods, traj, '-', linewidth=500*traj_weight[i], alpha=0.3)
    ax.plot(periods, oil_prices.T.mean(axis=0), '-k', linewidth=5)
    ax.set_ylim([0, 200])
    ax.set_xlim([2021, 2045])
    fig.show()

    # show how trajectories are clustered into representative paths
    # (choose the sample paths by inspecting oil_price_traj)
    fig = plt.figure()
    ax = fig.gca()
    for t_num in [20, 16]:
        sample_traj = oil_price_traj[t_num]
        traj_constituents = oil_prices.T[cluster_id==t_num, :]
        for i, traj in enumerate(traj_constituents):
            # plot each row as a separate series, with appropriate width and alpha
            ax.plot(periods, traj, '-', linewidth=1, alpha=0.3)
        ax.plot(periods, sample_traj, '-k', linewidth=5*500*traj_weight[t_num])
    #ax.set_ylim([0, 60])
    ax.set_xlim([2021, 2045])
    fig.show()

    # inspect the cluster trajectories and the first 1000 price series in 3D
    # to see if they are a good match
    fig = plt.figure()
    ax = fig.gca(projection='3d')
    for i in range(oil_price_traj.shape[0]):
        # plot each row as a separate series, with appropriate width and alpha
        ax.plot(periods, (oil_price_traj[i]), (gas_price_traj[i]), '-', linewidth=500*traj_weight[i], alpha=0.3)
    ax.plot(periods, np.log10(oil_prices.T.mean(axis=0)), np.log10(gas_prices.T.mean(axis=0)), '-k', linewidth=5)
    ax.set_ylim([0, 50])
    ax.set_zlim([0, 50])
    fig.show()

    fig = plt.figure()
    ax = fig.gca(projection='3d')
    for i in range(1000):
        # plot each row as a separate series, with appropriate width and alpha
        ax.plot(periods, (oil_prices.T[i]), (gas_prices.T[i]), '-', linewidth=1, alpha=0.3)
    ax.set_ylim([0, 50])
    ax.set_zlim([0, 50])
    fig.show()

# print "oil_prices"
# print oil_prices
# print "gas_prices"
# print gas_prices

# dictionary, with an array for each fuel supply tier, with 1 row for each year, 1 col for each scenario
fuel_prices = {
    (fuel, tier): om * oil_price_traj.T + gm * gas_price_traj.T + a 
        for (fuel, tier), (om, gm, a) in price_factors.iteritems()
}

# find indices of index columns
index_i = [headers.index(c) for c in ['regional_fuel_market', 'period', 'tier']]
# find indices of columns that should be omitted from the .dat file
# (these shouldn't be in the .tab file either, but they don't cause trouble there)
drop_i = [headers.index(c) for c in ['fuel']]
# helper function to sort a row of data from the .tab file into a suitable order for the .dat file
# (index columns must be shifted to the start of the row)
sort_cols = lambda r: (
    [r[i] for i in index_i] + [c for (i, c) in enumerate(r) if i not in index_i + drop_i]
)
# translate column names to the correct form for the model 
# (runph doesn't translate like our normal data loading functions)
col_name_dict = dict(zip(
        ["unit_cost", "max_avail_at_cost", "fixed_cost"],
        ["rfm_supply_tier_cost", "rfm_supply_tier_limit", "rfm_supply_tier_fixed_cost"]
    ))
translate_cols = lambda r: [col_name_dict.get(c, c) for c in r]

# write the data files for each scenario, using scenario-specific prices where available
for s in range(n_trajectories):
    fuel_data = []
    for row in standard_fuel_costs:
        # columns: 
        # regional_fuel_market, fuel, period, tier, unit_cost, max_avail_at_cost, fixed_cost
        r = list(row)   # copy the existing list before changing, or convert tuple to list
        if (r[1], r[3]) in fuel_prices:
            r[4] = str(fuel_prices[r[1], r[3]][periods.index(float(r[2])), s])
        fuel_data.append(r)
    with open(
        os.path.join(pha_dir, "fuel_supply_curves_{s}.dat".format(s=str(s).zfill(n_digits))),
        "w"
    ) as f:
        # omit headers for index cols
        f.write('param: ' + '\t'.join(translate_cols(sort_cols(headers)[3:])) + ' :=\n')
        for r in fuel_data:
            f.write('\t'.join(sort_cols(r)) + '\n')
        f.write(';\n')

# write all values to a .tsv file for graphing (not really used anymore)
with open(os.path.join(pha_dir, "fuel_supply_costs.tsv"), "w") as f:
    f.write("fuel\tyear\t" + "\t".join("price_per_mmbtu_"+str(s).zfill(n_digits) for s in range(n_trajectories)) + '\n')
    f.write('weight\t\t' + '\t'.join(traj_weight.astype(str).tolist()) + '\n')
    for fuel, tier in sorted(fuel_prices.keys()):
        for i, year in enumerate(periods):
            f.write('\t'.join([fuel, str(year)] + fuel_prices[fuel, tier][i].astype(str).tolist()) + '\n')

# write mean values to a tab file, dat file and tsv file
traj_weight_row = traj_weight[np.newaxis, :]
fuel_data = []
for row in standard_fuel_costs:
    # columns: 
    # regional_fuel_market, fuel, period, tier, unit_cost, max_avail_at_cost, fixed_cost
    r = list(row)   # copy the existing list before changing, or convert tuple to list
    if (r[1], r[3]) in fuel_prices:
        traj_prices = fuel_prices[r[1], r[3]][periods.index(float(r[2])), :]
        mean_price = np.sum(traj_prices * traj_weight_row, axis=1)[0]
        r[4] = str(mean_price)
    fuel_data.append(r)
with open(os.path.join(pha_mean_dir, "fuel_supply_curves.tab"), "w") as f:
    f.write('\t'.join(headers)+'\n')
    f.writelines('\t'.join(r)+'\n' for r in fuel_data)
with open(
    os.path.join(pha_mean_dir, "fuel_supply_curves_{s}.dat".format(s='0'.zfill(n_digits))),
    "w"
) as f:
    # omit headers for index cols
    f.write('param: ' + '\t'.join(translate_cols(sort_cols(headers)[3:])) + ' :=\n')
    f.writelines('\t'.join(sort_cols(r))+'\n' for r in fuel_data)
    f.write(';\n')

# write the weights for each scenario
with open(os.path.join(pha_dir, "scenario_weights.tsv"), "w") as f:
    f.write('scenario\tweight\n')
    f.writelines('{}\t{}\n'.format(i, w) for i, w in enumerate(traj_weight))
